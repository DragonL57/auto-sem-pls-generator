"""
Results exporter for Excel and other formats.
"""

import os
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Union
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..core.exceptions import ExportError


class ResultsExporter:
    """Exports results to Excel and other formats."""
    
    def __init__(self, config, output_dir: str):
        """
        Initialize results exporter.
        
        Args:
            config: Configuration object
            output_dir: Output directory
        """
        self.config = config
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_excel(self, data: pd.DataFrame, 
                    validation_results: Optional[Dict[str, Any]] = None,
                    parameters: Optional[np.ndarray] = None,
                    score: Optional[float] = None,
                    reason: Optional[str] = None) -> str:
        """
        Export results to Excel file.
        
        Args:
            data: Generated data
            validation_results: Validation results
            parameters: Best parameters
            score: Best score
            reason: Best reason
            
        Returns:
            Path to exported Excel file
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sem_results_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add data sheet
            self._add_data_sheet(wb, data)
            
            # Add validation results sheet
            if validation_results:
                self._add_validation_sheet(wb, validation_results)
            
            # Add configuration sheet
            self._add_configuration_sheet(wb)
            
            # Add parameters sheet
            if parameters is not None:
                self._add_parameters_sheet(wb, parameters, score, reason)
            
            # Add summary sheet
            self._add_summary_sheet(wb, data, validation_results, score)
            
            # Save workbook
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            raise ExportError(f"Failed to export Excel file: {str(e)}")
    
    def _add_data_sheet(self, wb: openpyxl.Workbook, data: pd.DataFrame):
        """Add data sheet to workbook."""
        ws = wb.create_sheet("Generated Data")
        
        # Write data
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_validation_sheet(self, wb: openpyxl.Workbook, validation_results: Dict[str, Any]):
        """Add validation results sheet."""
        ws = wb.create_sheet("Validation Results")
        
        # Write validation summary
        self._write_section_header(ws, "Validation Summary", 1)
        
        summary_data = [
            ["Overall Validity", "Yes" if validation_results.get('overall_validity', False) else "No"],
            ["Number of Recommendations", len(validation_results.get('recommendations', []))],
            ["Number of Warnings", len(validation_results.get('warnings', []))],
            ["Number of Errors", len(validation_results.get('errors', []))]
        ]
        
        for i, (key, value) in enumerate(summary_data, start=3):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value)
        
        # Write recommendations
        row_offset = 8
        self._write_section_header(ws, "Recommendations", row_offset)
        
        recommendations = validation_results.get('recommendations', [])
        for i, rec in enumerate(recommendations, start=row_offset + 2):
            ws.cell(row=i, column=1, value=f"• {rec}")
        
        # Write warnings
        row_offset += len(recommendations) + 4
        self._write_section_header(ws, "Warnings", row_offset)
        
        warnings = validation_results.get('warnings', [])
        for i, warning in enumerate(warnings, start=row_offset + 2):
            ws.cell(row=i, column=1, value=f"⚠ {warning}")
        
        # Write errors
        row_offset += len(warnings) + 4
        self._write_section_header(ws, "Errors", row_offset)
        
        errors = validation_results.get('errors', [])
        for i, error in enumerate(errors, start=row_offset + 2):
            ws.cell(row=i, column=1, value=f"❌ {error}")
        
        # Style the sheet
        self._style_validation_sheet(ws)
    
    def _add_configuration_sheet(self, wb: openpyxl.Workbook):
        """Add configuration sheet."""
        ws = wb.create_sheet("Configuration")
        
        # Write configuration summary
        self._write_section_header(ws, "Model Configuration", 1)
        
        config_data = [
            ["Number of Factors", self.config.n_latent_factors],
            ["Factor Names", ", ".join(self.config.latent_factor_names)],
            ["Number of Observations", self.config.num_observations],
            ["Number of Regression Models", len(self.config.regression_models)]
        ]
        
        for i, (key, value) in enumerate(config_data, start=3):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value)
        
        # Write factors configuration
        row_offset = 9
        self._write_section_header(ws, "Factors Configuration", row_offset)
        
        for i, (factor_name, factor_config) in enumerate(self.config.factors_config.items()):
            ws.cell(row=row_offset + 2 + i, column=1, value=factor_name)
            ws.cell(row=row_offset + 2 + i, column=2, value=", ".join(factor_config.original_items))
        
        # Write regression models
        row_offset += len(self.config.factors_config) + 5
        self._write_section_header(ws, "Regression Models", row_offset)
        
        for i, model in enumerate(self.config.regression_models):
            ws.cell(row=row_offset + 2 + i, column=1, value=f"{model.dependent} ~ {', '.join(model.independent)}")
        
        # Style the sheet
        self._style_configuration_sheet(ws)
    
    def _add_parameters_sheet(self, wb: openpyxl.Workbook, parameters: np.ndarray, 
                            score: float, reason: str):
        """Add parameters sheet."""
        ws = wb.create_sheet("Optimization Parameters")
        
        # Write optimization results
        self._write_section_header(ws, "Optimization Results", 1)
        
        results_data = [
            ["Best Score", f"{score:.3f}"],
            ["Reason", reason],
            ["Total Parameters", len(parameters)]
        ]
        
        for i, (key, value) in enumerate(results_data, start=3):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value)
        
        # Write parameter values
        row_offset = 8
        self._write_section_header(ws, "Parameter Values", row_offset)
        
        # Latent correlation values
        latent_cor_values = parameters[:self.config.n_latent_cor_values]
        ws.cell(row=row_offset + 2, column=1, value="Latent Correlation Values")
        for i, val in enumerate(latent_cor_values):
            ws.cell(row=row_offset + 3 + i, column=1, value=f"Correlation {i+1}")
            ws.cell(row=row_offset + 3 + i, column=2, value=f"{val:.3f}")
        
        # Error and loading strength
        error_strength = parameters[self.config.n_latent_cor_values]
        loading_strength = parameters[self.config.n_latent_cor_values + 1]
        
        param_row = row_offset + 3 + len(latent_cor_values) + 2
        ws.cell(row=param_row, column=1, value="Error Strength")
        ws.cell(row=param_row, column=2, value=f"{error_strength:.3f}")
        
        ws.cell(row=param_row + 1, column=1, value="Loading Strength")
        ws.cell(row=param_row + 1, column=2, value=f"{loading_strength:.3f}")
        
        # Style the sheet
        self._style_parameters_sheet(ws)
    
    def _add_summary_sheet(self, wb: openpyxl.Workbook, data: pd.DataFrame, 
                          validation_results: Optional[Dict[str, Any]], score: float):
        """Add summary sheet."""
        ws = wb.create_sheet("Summary")
        
        # Write summary information
        self._write_section_header(ws, "Analysis Summary", 1)
        
        summary_data = [
            ["Dataset Size", f"{len(data)} rows × {len(data.columns)} columns"],
            ["Optimization Score", f"{score:.3f}" if score else "N/A"],
            ["Generation Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Model Type", "SEM/PLS Synthetic Data Generation"]
        ]
        
        for i, (key, value) in enumerate(summary_data, start=3):
            ws.cell(row=i, column=1, value=key)
            ws.cell(row=i, column=2, value=value)
        
        # Write data statistics
        row_offset = 9
        self._write_section_header(ws, "Data Statistics", row_offset)
        
        # Calculate basic statistics
        numeric_data = data.select_dtypes(include=[np.number])
        stats_data = {
            "Mean": numeric_data.mean(),
            "Std Dev": numeric_data.std(),
            "Min": numeric_data.min(),
            "Max": numeric_data.max()
        }
        
        stats_df = pd.DataFrame(stats_data).T
        
        for i, (col, values) in enumerate(stats_df.items(), start=row_offset + 2):
            ws.cell(row=row_offset + 1, column=i + 1, value=col)
            for j, value in enumerate(values):
                ws.cell(row=row_offset + 2 + j, column=i + 1, value=f"{value:.3f}")
        
        # Write validation summary if available
        if validation_results:
            row_offset += len(stats_df) + 6
            self._write_section_header(ws, "Validation Summary", row_offset)
            
            validation_summary = [
                ["Overall Validity", "Pass" if validation_results.get('overall_validity', False) else "Fail"],
                ["Recommendations", len(validation_results.get('recommendations', []))],
                ["Warnings", len(validation_results.get('warnings', []))],
                ["Errors", len(validation_results.get('errors', []))]
            ]
            
            for i, (key, value) in enumerate(validation_summary, start=row_offset + 2):
                ws.cell(row=i, column=1, value=key)
                ws.cell(row=i, column=2, value=value)
        
        # Style the sheet
        self._style_summary_sheet(ws)
    
    def _write_section_header(self, ws: openpyxl.Worksheet, title: str, row: int):
        """Write a section header."""
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    def _style_validation_sheet(self, ws: openpyxl.Worksheet):
        """Style validation sheet."""
        # Add borders to cells
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
    
    def _style_configuration_sheet(self, ws: openpyxl.Worksheet):
        """Style configuration sheet."""
        # Add borders and alignment
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def _style_parameters_sheet(self, ws: openpyxl.Worksheet):
        """Style parameters sheet."""
        # Add borders and number formatting
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
                    
                    # Format numeric values
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = '0.000'
    
    def _style_summary_sheet(self, ws: openpyxl.Worksheet):
        """Style summary sheet."""
        # Add borders and formatting
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def export_json(self, results: Dict[str, Any], filename: Optional[str] = None) -> str:
        """
        Export results to JSON file.
        
        Args:
            results: Results dictionary
            filename: Optional filename
            
        Returns:
            Path to exported JSON file
        """
        import json
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sem_results_{timestamp}.json"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False, default=str)
            
            return filepath
            
        except Exception as e:
            raise ExportError(f"Failed to export JSON file: {str(e)}")
    
    def export_validation_report(self, validation_results: Dict[str, Any], 
                               filename: Optional[str] = None) -> str:
        """
        Export validation report to text file.
        
        Args:
            validation_results: Validation results
            filename: Optional filename
            
        Returns:
            Path to exported report file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"validation_report_{timestamp}.txt"
        
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("SEM/PLS Data Generation Validation Report\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                # Overall validity
                f.write("OVERALL VALIDITY\n")
                f.write("-" * 20 + "\n")
                f.write(f"Status: {'PASS' if validation_results.get('overall_validity', False) else 'FAIL'}\n\n")
                
                # Recommendations
                f.write("RECOMMENDATIONS\n")
                f.write("-" * 20 + "\n")
                for rec in validation_results.get('recommendations', []):
                    f.write(f"• {rec}\n")
                f.write("\n")
                
                # Warnings
                f.write("WARNINGS\n")
                f.write("-" * 20 + "\n")
                for warning in validation_results.get('warnings', []):
                    f.write(f"⚠ {warning}\n")
                f.write("\n")
                
                # Errors
                f.write("ERRORS\n")
                f.write("-" * 20 + "\n")
                for error in validation_results.get('errors', []):
                    f.write(f"❌ {error}\n")
                f.write("\n")
            
            return filepath
            
        except Exception as e:
            raise ExportError(f"Failed to export validation report: {str(e)}")